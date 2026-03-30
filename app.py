import streamlit as st
import io
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

st.set_page_config(page_title="Highlight Missing vs GL", layout="wide")
st.title("Highlight Missing Transactions vs GL")
st.write(
    "Upload the **Register** (bank transactions) and the **GL** (general ledger). "
    "Transactions in the register that are not found in the GL will be highlighted "
    "in orange. Matching is by **amount and date** (within a 5-day window). "
    "Transfers ending in **7459** are ignored."
)

ORANGE_FILL = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
TRANSFER_7459_KEYWORD = "7459"
DATE_TOLERANCE_DAYS = 5


def parse_date(value):
    """Parse a date value from either a datetime object or a string like '3/18/2026'."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    s = str(value).strip()
    for fmt in ("%m/%d/%Y", "%Y-%m-%d", "%m/%d/%y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def extract_gl_amounts(wb):
    """Extract (date, amount) pairs from the GL file.

    GL layout (columns by index):
      K (11) = Date, U (21) = Debit, W (23) = Credit
    In accounting for a cash/bank account:
      GL Debit  = money coming IN  (matches register credits)
      GL Credit = money going OUT  (matches register debits)

    Returns two lists of (date, amount) tuples.
    """
    ws = wb.active
    gl_debits = []   # money in  -> matches register credits
    gl_credits = []  # money out -> matches register debits

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        date_val = row[10].value if len(row) > 10 else None   # Column K
        debit_val = row[20].value if len(row) > 20 else None  # Column U
        credit_val = row[22].value if len(row) > 22 else None # Column W

        if date_val is None or date_val == "Date":
            continue

        dt = parse_date(date_val)
        if dt is None:
            continue

        if debit_val is not None:
            try:
                gl_debits.append((dt, round(float(debit_val), 2)))
            except (ValueError, TypeError):
                pass
        if credit_val is not None:
            try:
                gl_credits.append((dt, round(float(credit_val), 2)))
            except (ValueError, TypeError):
                pass

    return gl_debits, gl_credits


def is_transfer_7459(description):
    """Check if a transaction description refers to a transfer ending in 7459."""
    if description is None:
        return False
    return TRANSFER_7459_KEYWORD in str(description)


def find_match(reg_date, reg_amount, pool):
    """Find and remove a matching (date, amount) entry from the pool.

    A match requires the same amount and a date within DATE_TOLERANCE_DAYS.
    If multiple matches exist, prefer the closest date.
    Returns True if a match was found and consumed.
    """
    best_idx = None
    best_diff = None

    for i, (gl_date, gl_amount) in enumerate(pool):
        if gl_amount == reg_amount:
            if reg_date is not None and gl_date is not None:
                diff = abs((reg_date - gl_date).days)
                if diff <= DATE_TOLERANCE_DAYS:
                    if best_diff is None or diff < best_diff:
                        best_idx = i
                        best_diff = diff
            elif reg_date is None or gl_date is None:
                # If either date is missing, match on amount alone
                if best_idx is None:
                    best_idx = i
                    best_diff = 0

    if best_idx is not None:
        pool.pop(best_idx)
        return True
    return False


def find_missing_and_highlight(register_wb, gl_debits, gl_credits):
    """Find register transactions missing from the GL and highlight them orange.

    Returns the modified workbook and counts of missing / total / ignored.
    """
    ws = register_wb.active

    # Build consumable pools
    gl_credit_pool = list(gl_credits)  # for matching register debits (out)
    gl_debit_pool = list(gl_debits)    # for matching register credits (in)

    total = 0
    missing = 0
    ignored = 0

    for row_idx in range(2, ws.max_row + 1):
        date_cell = ws.cell(row=row_idx, column=1).value
        if date_cell is None:
            continue
        date_str = str(date_cell)
        # Skip non-transaction rows
        if any(kw in date_str for kw in ["TOTALS", "Total", "Beginning", "Ending", "balance"]):
            continue

        reg_date = parse_date(date_cell)
        description = ws.cell(row=row_idx, column=3).value or ""
        debit_val = ws.cell(row=row_idx, column=4).value   # Debits (Out)
        credit_val = ws.cell(row=row_idx, column=5).value   # Credits (In)

        # Skip transfers ending in 7459
        if is_transfer_7459(description):
            ignored += 1
            continue

        total += 1
        is_missing = False

        # Check debit (money out): match against GL credits
        if debit_val is not None:
            try:
                amt = round(float(debit_val), 2)
                if amt != 0:
                    if not find_match(reg_date, amt, gl_credit_pool):
                        is_missing = True
            except (ValueError, TypeError):
                pass

        # Check credit (money in): match against GL debits
        if credit_val is not None:
            try:
                amt = round(float(credit_val), 2)
                if amt != 0:
                    if not find_match(reg_date, amt, gl_debit_pool):
                        is_missing = True
            except (ValueError, TypeError):
                pass

        if is_missing:
            missing += 1
            # Highlight the entire row in orange
            for col_idx in range(1, ws.max_column + 1):
                ws.cell(row=row_idx, column=col_idx).fill = ORANGE_FILL

    return register_wb, total, missing, ignored


# --------------- Streamlit UI ---------------

col1, col2 = st.columns(2)
with col1:
    register_file = st.file_uploader("Upload Register (Excel)", type=["xlsx", "xls"])
with col2:
    gl_file = st.file_uploader("Upload GL (Excel)", type=["xlsx", "xls"])

if register_file and gl_file:
    if st.button("Compare & Highlight Missing", type="primary"):
        with st.spinner("Loading GL..."):
            gl_wb = load_workbook(io.BytesIO(gl_file.read()), data_only=True)
            gl_debits, gl_credits = extract_gl_amounts(gl_wb)

        st.info(f"GL loaded: **{len(gl_debits)}** debit entries, **{len(gl_credits)}** credit entries.")

        with st.spinner("Comparing register against GL..."):
            register_wb = load_workbook(io.BytesIO(register_file.read()))
            register_wb, total, missing, ignored = find_missing_and_highlight(
                register_wb, gl_debits, gl_credits
            )

        # Summary
        matched = total - missing
        st.success(
            f"Done! **{total}** transactions checked, **{matched}** matched, "
            f"**{missing}** missing (highlighted orange), **{ignored}** transfers to 7459 ignored."
        )

        if missing > 0:
            st.warning(f"{missing} transaction(s) highlighted in orange are missing from the GL.")

        # Save to bytes for download
        output = io.BytesIO()
        register_wb.save(output)
        output.seek(0)

        st.download_button(
            label="Download Highlighted Register",
            data=output.getvalue(),
            file_name="register_highlighted.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.spreadsheetml",
            type="primary",
        )
